import io
import sqlite3
import time
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
FILES = {
    "2G": "1zwa8F_WrJS9LXArcNJqmemn7FK96Ycnu",
    "3G": "1HCI7IuWjMle50E-TRAz1cbo-yZLaUPF-",
    "4G": "13q7CDdLC0Hy4lmgyu9PY-EA-hBgDGFpj",
    "5G": "1Ff7NnCsDQl0YdbDxvEEa82rrFtYzojVB",
}
DOWNLOAD_URLS = (
    "https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t",
    "https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx",
)
FIELDS = [
    "site", "uf", "banda", "azimuth", "bcch", "psc", "pci", "bandwidth",
    "cidade", "bairro", "endereco", "earfcn", "latitude", "longitude", "mimo",
    "dl_earfcn", "dl_uarfcn",
]
COLUMNS = {
    "site": "[P]SITE", "uf": "[P]UF", "banda": "[P]BANDA_OPERACAO",
    "azimuth": "[P]AZIMUTH", "bcch": "[P]BCCH", "psc": "[P]PSC",
    "pci": "[P]PCI", "bandwidth": "[P]BANDWIDTH", "cidade": "[P]CIDADE",
    "bairro": "[P]BAIRRO", "endereco": "[P]ENDERECO", "latitude": "[P]LATITUDE",
    "longitude": "[P]LONGITUDE", "mimo": "[P]MIMO", "dl_earfcn": "[P]DL_EARFCN",
    "dl_uarfcn": "[P]DL_UARFCN",
}


def text(value):
    if value is None:
        return ""
    value = str(value).strip()
    return "" if value.lower() == "nan" else value


def download_workbook(file_id):
    headers = {"User-Agent": "Mozilla/5.0"}
    last_error = None
    for url_template in DOWNLOAD_URLS:
        url = url_template.format(file_id=file_id)
        for attempt in range(3):
            try:
                request = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(request, timeout=60) as response:
                    return response.read()
            except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as error:
                last_error = error
                time.sleep(2 * (attempt + 1))
    raise last_error


def build():
    db_path = ROOT / "rf_cache.db"
    if db_path.exists():
        db_path.unlink()
    connection = sqlite3.connect(db_path)
    connection.execute("CREATE TABLE rf (id INTEGER PRIMARY KEY, site TEXT, uf TEXT, tech TEXT, banda TEXT, azimuth TEXT, bcch TEXT, psc TEXT, pci TEXT, bandwidth TEXT, cidade TEXT, bairro TEXT, endereco TEXT, earfcn TEXT, latitude TEXT, longitude TEXT, mimo TEXT)")
    insert = "INSERT INTO rf (site,uf,tech,banda,azimuth,bcch,psc,pci,bandwidth,cidade,bairro,endereco,earfcn,latitude,longitude,mimo) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
    for tech, file_id in FILES.items():
        workbook = openpyxl.load_workbook(io.BytesIO(download_workbook(file_id)), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [text(value) for value in next(iterator)]
        indexes = {name: index for index, name in enumerate(headers)}
        for row in iterator:
            values = {field: text(row[indexes[COLUMNS[field]]]) if field in COLUMNS and COLUMNS[field] in indexes and indexes[COLUMNS[field]] < len(row) else "" for field in FIELDS}
            earfcn = values["dl_uarfcn"] if tech == "3G" else values["dl_earfcn"] if tech in ("4G", "5G") else ""
            connection.execute(insert, (values["site"].upper(), values["uf"].upper(), tech, values["banda"], values["azimuth"], values["bcch"] if tech == "2G" else "", values["psc"] if tech == "3G" else "", values["pci"] if tech in ("4G", "5G") else "", values["bandwidth"] if tech == "4G" else "", values["cidade"], values["bairro"], values["endereco"], earfcn, values["latitude"], values["longitude"], values["mimo"] if tech == "4G" else ""))
        workbook.close()
    connection.execute("CREATE INDEX ix_rf_site_uf ON rf(site, uf)")
    connection.commit()
    connection.close()


if __name__ == "__main__":
    build()
